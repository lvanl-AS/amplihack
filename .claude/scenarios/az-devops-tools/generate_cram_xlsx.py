#!/usr/bin/env python3
"""Generate a CRAM (Change Risk Assessment Matrix) Excel workbook.

Takes scores as JSON input and produces a .xlsx file matching the official
Alaska Airlines CRAM template — same layout, formulas, and color coding.

Usage:
    # From JSON argument
    python generate_cram_xlsx.py --scores '{"system_criticality": 2, ...}' -o CRAM-output.xlsx

    # From stdin
    echo '{"system_criticality": 2, ...}' | python generate_cram_xlsx.py -o CRAM-output.xlsx

    # From file
    python generate_cram_xlsx.py --scores-file scores.json -o CRAM-output.xlsx

Score JSON format:
    {
        "system_criticality": 1-3,
        "customer_impact": 1-3,
        "regulatory_compliance": 1-3,
        "change_complexity": 1-3,
        "testing_coverage": 1-3,
        "rollback_feasibility": 1-3,
        "change_window": 1-3,
        "operational_readiness": 1-3,
        "vendor_dependency": 1-3,
        "secure_coding": 1-3,
        "resilience_recovery": 1-3
    }
"""

import argparse
import json
import math
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Alignment,
        Border,
        Font,
        PatternFill,
        Side,
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# ---------------------------------------------------------------------------
# Constants — rubric text from official CRAM template
# ---------------------------------------------------------------------------

SUBCRITERIA = [
    {
        "key": "system_criticality",
        "name": "System Criticality",
        "category": "business_impact",
        "low": "Tier 3 or Tier 4 systems; impacts a single department",
        "medium": "Tier 2 systems; impacts a single site or multiple departments",
        "high": "Tier 0 or Tier 1 systems OR impacts multiple sites / critical operations",
    },
    {
        "key": "customer_impact",
        "name": "Customer Impact",
        "category": "business_impact",
        "low": "No potential customer impact; internal-only change",
        "medium": "Limited potential customer impact; minor service degradation or short outage",
        "high": "Significant potential customer impact; major outage or service disruption",
    },
    {
        "key": "regulatory_compliance",
        "name": "Regulatory Compliance",
        "category": "business_impact",
        "low": "No compliance implications",
        "medium": "Minor compliance considerations; documented and approved",
        "high": "High compliance risk; regulatory breach possible if change fails",
    },
    {
        "key": "change_complexity",
        "name": "Change Complexity",
        "category": "execution_risk",
        "low": "Simple change; 1\u20132 teams, single system, few steps, minimal risk, limited scope, predictable outcome",
        "medium": "Moderate complexity; 3\u20134 teams, one complex system or minor integration",
        "high": "High complexity; 5+ teams, multiple systems, deep integration, unpredictable outcomes",
    },
    {
        "key": "testing_coverage",
        "name": "Testing Coverage",
        "category": "execution_risk",
        "low": "Fully tested, evidence available, low risk",
        "medium": "Partially tested, failed test cases, risk acknowledged, acceptable with oversight",
        "high": "Untested or failed, no mitigation; high risk if deployed",
    },
    {
        "key": "rollback_feasibility",
        "name": "Rollback Feasibility",
        "category": "execution_risk",
        "low": "Easy rollback; < 30 minutes",
        "medium": "Moderate rollback; 30 min to 2 hours",
        "high": "Difficult rollback; > 2 hours or requires major recovery effort",
    },
    {
        "key": "change_window",
        "name": "Change Window",
        "category": "operational_risk",
        "low": "No outage expected or within the Approved Maintenance Window",
        "medium": "Outage occuring during Off-Hours (22:30 - 04:49) or during the weekend Soft-Freeze",
        "high": "Outage during Peak / Business Hours / Pre-Flight / Global Ops (05:00\u201317:00) or during a Hard Freeze",
    },
    {
        "key": "operational_readiness",
        "name": "Operational Readiness",
        "category": "operational_risk",
        "low": "All documentation, observability, approvals, and resources confirmed; support teams ready",
        "medium": "Some gaps in documentation, observability, or readiness; mitigations in place",
        "high": "Significant gaps; unclear ownership or missing approvals",
    },
    {
        "key": "vendor_dependency",
        "name": "Vendor Dependency",
        "category": "operational_risk",
        "low": "No vendor involvement; fully internal",
        "medium": "Vendor involvement for minor tasks or advisory",
        "high": "Vendor dependency for critical tasks; delays or coordination risks",
    },
    {
        "key": "secure_coding",
        "name": "Secure Coding",
        "category": "operational_risk",
        "low": "Standard Quality Gate (or higher) is used; No Critical or High vulnerabilities are present",
        "medium": "Security Plus or Security Quality Gate is used; Unresolved security hotspots are present",
        "high": "No Quality Gate is configured; Critical or High vulnerabilities exist, and security hotspots remain unresolved",
    },
    {
        "key": "resilience_recovery",
        "name": "Resilience & Recovery (RR)",
        "category": "operational_risk",
        "low": "RR plan fully automated, tested successfully, and evidence demonstrating less than 1 hour total elapsed time for recovery prior to outside validation",
        "medium": "RR plan tested and lacking full automation or with issues remediated without re-testing; Recovery process is < 2 hours total elapsed time for recovery prior to outside validation",
        "high": "No RR testing evidence; recovery plan unverified and/or recovery process is > 2 hours total elapsed time for recovery prior to outside validation",
    },
]

# Category definitions with their weighting rules
CATEGORIES = {
    "business_impact": {
        "label": "Business Impact (Weight = 3)",
        "sublabel": "Highest Score X 3",
        "description": "Biggest driver; What could this break?",
        "weight": 3,
        "method": "highest",
    },
    "execution_risk": {
        "label": "Execution Risk (Weight = 2)",
        "sublabel": "Average Score X 2",
        "description": "Second driver; How hard is it to do safely?",
        "weight": 2,
        "method": "average",
    },
    "operational_risk": {
        "label": "Operational Risk (Weight = 1)",
        "sublabel": "Average Score X 1",
        "description": "Still important, but lighter weight; Are we truly ready to run it now?",
        "weight": 1,
        "method": "average",
    },
}

# ---------------------------------------------------------------------------
# Color scheme — matched from official template and team examples
# ---------------------------------------------------------------------------

# Score cell colors (1=green, 2=yellow, 3=red)
SCORE_FILLS = {
    1: PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),  # light green
    2: PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),  # light yellow
    3: PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),  # light red
}

SCORE_FONTS = {
    1: Font(color="006100", bold=True),   # dark green
    2: Font(color="9C5700", bold=True),   # dark yellow/brown
    3: Font(color="9C0006", bold=True),   # dark red
}

# Legend colors
LEGEND_FILLS = {
    "low": PatternFill(start_color="92D050", end_color="92D050", fill_type="solid"),
    "medium": PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"),
    "high": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
}

# Header fills
HEADER_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")  # light blue
CATEGORY_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # light gray
TITLE_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # blue

# Borders
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


# ---------------------------------------------------------------------------
# Score calculation
# ---------------------------------------------------------------------------

def calculate_weighted_scores(scores: dict) -> dict:
    """Calculate weighted scores from raw subcriteria scores."""
    category_scores = {"business_impact": [], "execution_risk": [], "operational_risk": []}

    for sc in SUBCRITERIA:
        val = scores.get(sc["key"], 0)
        category_scores[sc["category"]].append(val)

    result = {}
    for cat_key, cat_def in CATEGORIES.items():
        raw_scores = category_scores[cat_key]
        if cat_def["method"] == "highest":
            raw = max(raw_scores) if raw_scores else 0
        else:
            raw = sum(raw_scores) / len(raw_scores) if raw_scores else 0

        weighted = round(raw * cat_def["weight"])
        result[cat_key] = {"raw": raw, "weighted": weighted}

    result["total"] = sum(v["weighted"] for v in result.values() if isinstance(v, dict))

    if result["total"] <= 10:
        result["risk_level"] = "Low"
    elif result["total"] <= 15:
        result["risk_level"] = "Medium"
    else:
        result["risk_level"] = "High"

    return result


# ---------------------------------------------------------------------------
# Excel generation
# ---------------------------------------------------------------------------

def _apply_border(ws, row, col_start, col_end):
    """Apply thin border to a range of cells in a row."""
    for c in range(col_start, col_end + 1):
        ws.cell(row=row, column=c).border = THIN_BORDER


def _set_cell(ws, row, col, value, font=None, fill=None, alignment=None, border=True):
    """Set cell value with optional formatting."""
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = THIN_BORDER
    return cell


def generate_workbook(scores: dict) -> Workbook:
    """Generate a CRAM workbook from scores dict."""
    wb = Workbook()
    ws = wb.active
    ws.title = "CRAM"

    # Column widths (A=CRAM label, B=Criteria, C=Low, D=Medium, E=High, F=Score, G=Weighted)
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 40
    ws.column_dimensions["F"].width = 8
    ws.column_dimensions["G"].width = 12

    wrap = Alignment(wrap_text=True, vertical="center")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    center_bold = Font(bold=True, size=11)

    # ── Row 1: Title row ──
    row = 1
    _set_cell(ws, row, 1, "CRAM\n(Change Risk Assessment Matrix)",
              font=Font(bold=True, size=12), fill=TITLE_FILL,
              alignment=Alignment(wrap_text=True, vertical="center", horizontal="center"))
    ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="FFFFFF")

    _set_cell(ws, row, 2, "Criteria", font=Font(bold=True, color="FFFFFF"), fill=TITLE_FILL, alignment=center)

    for col, label in [(3, "Low Risk (1)\nSafe, routine changes, with low complexity and predictable outcomes"),
                        (4, "Medium Risk (2)\nModerately complex changes that require oversight, scheduling review, and leadership review"),
                        (5, "High Risk (3)\nSignificant risk of major incidents, requires senior leadership review, stringent controls, and strategic scheduling")]:
        _set_cell(ws, row, col, label, font=Font(bold=True, size=9, color="FFFFFF"),
                  fill=TITLE_FILL, alignment=center)

    _set_cell(ws, row, 6, "Score", font=Font(bold=True, color="FFFFFF"), fill=TITLE_FILL, alignment=center)
    _set_cell(ws, row, 7, "Weighted\nScore", font=Font(bold=True, color="FFFFFF"), fill=TITLE_FILL, alignment=center)

    ws.row_dimensions[row].height = 50

    # ── Subcriteria rows ──
    calc = calculate_weighted_scores(scores)

    # Track which rows belong to which category for the merged label column
    category_rows = {"business_impact": [], "execution_risk": [], "operational_risk": []}
    row = 2

    for sc in SUBCRITERIA:
        category_rows[sc["category"]].append(row)

        score_val = scores.get(sc["key"], 0)

        # B: Criteria name
        _set_cell(ws, row, 2, sc["name"], font=Font(bold=True, size=10), alignment=wrap)

        # C, D, E: Rubric text
        _set_cell(ws, row, 3, sc["low"], font=Font(size=9), alignment=wrap)
        _set_cell(ws, row, 4, sc["medium"], font=Font(size=9), alignment=wrap)
        _set_cell(ws, row, 5, sc["high"], font=Font(size=9), alignment=wrap)

        # F: Score with conditional color
        score_fill = SCORE_FILLS.get(score_val)
        score_font = SCORE_FONTS.get(score_val, Font(bold=True))
        _set_cell(ws, row, 6, score_val, font=score_font, fill=score_fill, alignment=center)

        # G: Weighted score (only on the first row of each category)
        # Leave blank for now, fill after merge
        _set_cell(ws, row, 7, None, alignment=center)

        ws.row_dimensions[row].height = 45
        row += 1

    # ── Category labels (column A) — merged cells ──
    for cat_key, cat_def in CATEGORIES.items():
        rows = category_rows[cat_key]
        if not rows:
            continue

        first_row, last_row = rows[0], rows[-1]
        label = f"{cat_def['label']}\n{cat_def['sublabel']}\n\n{cat_def['description']}"

        ws.merge_cells(start_row=first_row, start_column=1, end_row=last_row, end_column=1)
        cell = ws.cell(row=first_row, column=1, value=label)
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.fill = CATEGORY_FILL
        cell.border = THIN_BORDER

        # Apply border to all merged cells
        for r in rows:
            ws.cell(row=r, column=1).border = THIN_BORDER

        # Weighted score — merged in column G for this category
        ws.merge_cells(start_row=first_row, start_column=7, end_row=last_row, end_column=7)
        weighted_val = calc[cat_key]["weighted"]

        # Color the weighted score based on risk contribution
        w_cell = ws.cell(row=first_row, column=7, value=weighted_val)
        w_cell.font = Font(bold=True, size=14)
        w_cell.alignment = center
        w_cell.border = THIN_BORDER
        for r in rows:
            ws.cell(row=r, column=7).border = THIN_BORDER

    # ── Legend and total section ──
    # Row: Legend label (col 1-4 blank, col 5 "Legend") + Total label + Total score
    legend_row_1 = row
    ws.merge_cells(start_row=legend_row_1, start_column=1, end_row=legend_row_1, end_column=4)

    _set_cell(ws, legend_row_1, 5, "Legend", font=Font(bold=True), alignment=center)
    _set_cell(ws, legend_row_1, 6, "Total\nWeighted\nScore ->",
              font=Font(bold=True, size=9), alignment=center)

    total = calc["total"]
    risk = calc["risk_level"]

    if risk == "Low":
        total_fill = LEGEND_FILLS["low"]
    elif risk == "Medium":
        total_fill = LEGEND_FILLS["medium"]
    else:
        total_fill = LEGEND_FILLS["high"]

    _set_cell(ws, legend_row_1, 7, total, font=Font(bold=True, size=16), fill=total_fill, alignment=center)
    ws.row_dimensions[legend_row_1].height = 40

    row = legend_row_1 + 1

    # Legend rows: Low, Medium, High
    for level, label, fill in [
        ("low", "Low Risk: 6-10", LEGEND_FILLS["low"]),
        ("medium", "Medium Risk: 11-15", LEGEND_FILLS["medium"]),
        ("high", "High Risk: 16-18", LEGEND_FILLS["high"]),
    ]:
        cell = ws.cell(row=row, column=5, value=label)
        cell.font = Font(bold=True, size=10, color="FFFFFF" if level == "high" else "000000")
        cell.fill = fill
        cell.alignment = center
        cell.border = THIN_BORDER
        row += 1

    # ── Print settings ──
    ws.sheet_properties.pageSetUpPr = None
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1

    return wb


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

SCORE_KEYS = [sc["key"] for sc in SUBCRITERIA]


def validate_scores(scores: dict) -> dict:
    """Validate and normalize score input."""
    result = {}
    for key in SCORE_KEYS:
        val = scores.get(key)
        if val is None:
            print(f"Warning: missing score for '{key}', defaulting to 0", file=sys.stderr)
            result[key] = 0
            continue
        try:
            val = int(val)
        except (TypeError, ValueError):
            print(f"Error: score for '{key}' must be 1-3, got '{val}'", file=sys.stderr)
            sys.exit(1)
        if val not in (1, 2, 3):
            print(f"Error: score for '{key}' must be 1-3, got {val}", file=sys.stderr)
            sys.exit(1)
        result[key] = val
    return result


def main():
    parser = argparse.ArgumentParser(description="Generate CRAM Excel workbook")
    parser.add_argument("-o", "--output", required=True, help="Output .xlsx file path")
    parser.add_argument("--scores", help="Scores as JSON string")
    parser.add_argument("--scores-file", help="Path to JSON file with scores")
    args = parser.parse_args()

    # Load scores
    if args.scores:
        scores = json.loads(args.scores)
    elif args.scores_file:
        scores = json.loads(Path(args.scores_file).read_text())
    elif not sys.stdin.isatty():
        scores = json.loads(sys.stdin.read())
    else:
        parser.error("Provide scores via --scores, --scores-file, or stdin")

    scores = validate_scores(scores)
    wb = generate_workbook(scores)

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))

    # Print summary to stderr, path to stdout
    calc = calculate_weighted_scores(scores)
    print(f"CRAM generated: {calc['risk_level']} Risk (score: {calc['total']})", file=sys.stderr)
    print(str(output_path))


if __name__ == "__main__":
    main()


__all__ = ["generate_workbook", "calculate_weighted_scores", "validate_scores", "SUBCRITERIA", "CATEGORIES"]
